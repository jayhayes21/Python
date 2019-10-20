from openpyxl import load_workbook
import datetime
from datetime import timedelta
import matplotlib.pyplot as plt

running_log = r"C:\Users\Jay Hayes\Desktop\Running Log.xlsx"
run_mileage = []
cooper_run_mileage = []
harris_run_mileage = []
run_week = []
start_date = datetime.datetime(2017, 8, 21)
# The 6th column is our running mileage row and we want to know when this is blank

def find_last_element():
    workbook = load_workbook(filename=running_log)
    sheet = workbook.active
    x = 1
    while True:
        if sheet.cell(row=x, column=6).value is None:
            return x-1
        x += 1
    workbook.close()

def plot_last_month(last_element):
    workbook = load_workbook(filename=running_log)
    sheet = workbook.active
    week_inc = timedelta(days=7)
    today = start_date + ((last_element-19-3)*week_inc)
    for x in range(4):
        row_number = last_element - (3 - x)
        run_mileage.append(sheet.cell(row=row_number, column=6).value)
        harris_run_mileage.append(sheet.cell(row=row_number, column=10).value)
        cooper_run_mileage.append(sheet.cell(row=row_number, column=11).value)
        run_week.append(today)
        today += week_inc
    workbook.close()
    plt.plot([str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10]],
             [run_mileage[0], run_mileage[1], run_mileage[2], run_mileage[3]], label='Me')
    plt.plot([str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10]],
             [run_mileage[0], run_mileage[1], run_mileage[2], run_mileage[3]], 'bo')

    plt.plot([str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10]],
             [harris_run_mileage[0], harris_run_mileage[1], harris_run_mileage[2], harris_run_mileage[3]], label='Harris')
    plt.plot([str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10]],
             [harris_run_mileage[0], harris_run_mileage[1], harris_run_mileage[2], harris_run_mileage[3]], 'yo')

    plt.plot([str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10]],
             [cooper_run_mileage[0], cooper_run_mileage[1], cooper_run_mileage[2], cooper_run_mileage[3]], label='Cooper')
    plt.plot([str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10]],
             [cooper_run_mileage[0], cooper_run_mileage[1], cooper_run_mileage[2], cooper_run_mileage[3]], 'go')
    plt.legend()
    plt.ylabel('Running Miles')
    plt.xlabel('Week of Runs')
    plt.title('Running Miles the past month')
    plt.show()


def plot_last_two_months(last_element):
    workbook = load_workbook(filename=running_log)
    sheet = workbook.active
    week_inc = timedelta(days=7)
    today = start_date + ((last_element-19-7)*week_inc)
    for x in range(8):
        row_number = last_element - (7 - x)
        run_mileage.append(sheet.cell(row=row_number, column=6).value)
        harris_run_mileage.append(sheet.cell(row=row_number, column=10).value)
        cooper_run_mileage.append(sheet.cell(row=row_number, column=11).value)
        run_week.append(today)
        today += week_inc
    workbook.close()
    plt.plot([str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10], str(run_week[4])[0:10], str(run_week[5])[0:10], str(run_week[6])[0:10], str(run_week[7])[0:10]],
             [run_mileage[0], run_mileage[1], run_mileage[2], run_mileage[3], run_mileage[4], run_mileage[5], run_mileage[6], run_mileage[7]], label='Me')
    plt.plot(
        [str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10], str(run_week[4])[0:10], str(run_week[5])[0:10],str(run_week[6])[0:10], str(run_week[7])[0:10]],
        [run_mileage[0], run_mileage[1], run_mileage[2], run_mileage[3], run_mileage[4], run_mileage[5], run_mileage[6],run_mileage[7]], 'bo')

    plt.plot([str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10],
              str(run_week[4])[0:10], str(run_week[5])[0:10], str(run_week[6])[0:10], str(run_week[7])[0:10]],
             [harris_run_mileage[0], harris_run_mileage[1], harris_run_mileage[2], harris_run_mileage[3], harris_run_mileage[4], harris_run_mileage[5],
              harris_run_mileage[6], harris_run_mileage[7]], label='Harris')
    plt.plot(
        [str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10],
         str(run_week[4])[0:10], str(run_week[5])[0:10], str(run_week[6])[0:10], str(run_week[7])[0:10]],
        [harris_run_mileage[0], harris_run_mileage[1], harris_run_mileage[2], harris_run_mileage[3], harris_run_mileage[4], harris_run_mileage[5], harris_run_mileage[6],
         harris_run_mileage[7]], 'yo')

    plt.plot([str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10],
              str(run_week[4])[0:10], str(run_week[5])[0:10], str(run_week[6])[0:10], str(run_week[7])[0:10]],
             [cooper_run_mileage[0], cooper_run_mileage[1], cooper_run_mileage[2], cooper_run_mileage[3], cooper_run_mileage[4], cooper_run_mileage[5],
              cooper_run_mileage[6], cooper_run_mileage[7]], label='Cooper')
    plt.plot(
        [str(run_week[0])[0:10], str(run_week[1])[0:10], str(run_week[2])[0:10], str(run_week[3])[0:10],
         str(run_week[4])[0:10], str(run_week[5])[0:10], str(run_week[6])[0:10], str(run_week[7])[0:10]],
        [cooper_run_mileage[0], cooper_run_mileage[1], cooper_run_mileage[2], cooper_run_mileage[3], cooper_run_mileage[4], cooper_run_mileage[5], cooper_run_mileage[6],
         cooper_run_mileage[7]], 'go')
    plt.legend()
    plt.ylabel('Running Miles')
    plt.xlabel('Week of Runs')
    plt.title('Running Miles the past 2 months')
    plt.show()